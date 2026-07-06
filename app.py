"""
app.py - DSD Route Planner Web Application
"""

import io
import os
import sys
import time
import threading
from datetime import date

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

ROOT = os.path.dirname(os.path.abspath(__file__))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import config
import route_planner as rp
import fleet_registry as fr

FLEET_FILE = "_setup/fleet.xlsx"

st.set_page_config(
    page_title="DSD Route Planner",
    page_icon="🚛",
    layout="wide",
    initial_sidebar_state="auto",
)

st.markdown("""
<style>
  .block-container { padding-top: 1.2rem; }
  .stButton > button { border-radius: 6px; }
  div[data-testid="metric-container"] {
    background: #f8f9fa; border-radius: 8px; padding: 8px;
  }
  /* Hide only the top-right action buttons, keep sidebar toggle intact */
  #MainMenu { visibility: hidden; }
  footer { visibility: hidden; }
  /* Target the deploy/share/star buttons specifically */
  .stDeployButton { display: none; }
  div[data-testid="stDecoration"] { display: none; }
</style>
""", unsafe_allow_html=True)

st.sidebar.title("🚛 DSD Route Planner")
st.sidebar.caption(f"📅 {date.today().strftime('%A, %d %b %Y')}")
st.sidebar.markdown("---")
page = st.sidebar.radio(
    "Navigate",
    ["🗺 Route Planning", "⚙️ Admin"],
    label_visibility="collapsed",
)
st.sidebar.markdown("---")
st.sidebar.caption("v2.0 · Skopje DSD")


def make_blank_template() -> bytes:
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(color="FFFFFF", bold=True)
    wb = openpyxl.Workbook()
    ws_o = wb.active
    ws_o.title = "Orders"
    ws_o.append(["customer_code", "customer_name", "cases", "kg"])
    for cell in ws_o[1]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    for col, w in zip(["A","B","C","D"], [18, 35, 10, 12]):
        ws_o.column_dimensions[col].width = w
    ws_o.freeze_panes = "A2"

    ws_v = wb.create_sheet("Vehicles")
    headers = ["vehicle_name", "vehicle_type", "capacity_kg",
               "zone", "available", "max_trips_per_day", "notes"]
    ws_v.append(headers)
    for cell in ws_v[1]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    for col_letter, w in zip(["A","B","C","D","E","F","G"], [20,12,14,22,12,18,25]):
        ws_v.column_dimensions[col_letter].width = w
    ws_v.append([])
    ws_v.append(["HOW TO FILL:"])
    ws_v.append(["vehicle_name    -> Your vehicle name/plate (e.g. DSD 1, F 3, Van 7)"])
    ws_v.append(["vehicle_type    -> Kamion | Furgon | Van"])
    ws_v.append(["capacity_kg     -> Usable kg per trip (e.g. 6000, 5200, 3200)"])
    ws_v.append(["zone            -> Same as vehicle_name for dedicated zone. Float for overflow."])
    ws_v.append(["available       -> TRUE = working today, FALSE = unavailable"])
    ws_v.append(["max_trips       -> 2 = normal day, 3 = heavy day"])
    ws_v.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def run_routing(file_bytes: bytes, mode: str):
    with open(config.TODAY_FILE, "wb") as f:
        f.write(file_bytes)
    stops_df    = rp.load_orders(config.TODAY_FILE, config.CUSTOMER_MASTER)
    vehicles_df = rp.load_vehicles(config.TODAY_FILE)
    zone_affinity = (mode == "smart")
    routes, unassigned, stops_df = rp.solve(stops_df, vehicles_df, zone_affinity=zone_affinity)
    zone_summary = None
    if zone_affinity:
        vehicle_zones = {}
        for _, row in vehicles_df.iterrows():
            z = str(row["zone"]).strip()
            if z.lower() not in ("", "nan", "float", "none"):
                vehicle_zones[row["vehicle_name"]] = z
        original_counts = {}
        for _, row in stops_df.iterrows():
            sz = str(row.get("zone", "")).strip()
            for vname, vzone in vehicle_zones.items():
                if vzone == sz:
                    original_counts[vname] = original_counts.get(vname, 0) + 1
                    break
        zone_summary = rp.build_zone_summary_from_routes(routes, original_counts)
    excel_buf = io.BytesIO()
    rp.write_excel(routes, stops_df, unassigned, excel_buf, zone_summary)
    excel_buf.seek(0)
    map_html = rp.generate_route_map(routes) if routes else None
    return excel_buf.getvalue(), routes, stops_df, unassigned, zone_summary, map_html


def run_routing_threaded(file_bytes: bytes, mode: str, result_holder: dict):
    """Wrapper that runs run_routing in a thread and stores result/error."""
    try:
        result_holder["result"] = run_routing(file_bytes, mode)
    except SystemExit as e:
        result_holder["error"] = str(e)
    except Exception as e:
        result_holder["error"] = str(e)


def check_zones_assigned(file_bytes: bytes) -> bool:
    try:
        xl  = pd.ExcelFile(io.BytesIO(file_bytes))
        veh = xl.parse("Vehicles")
        veh.columns = veh.columns.str.lower().str.strip()
        veh = veh.dropna(subset=["vehicle_name"])
        zones = veh["zone"].fillna("").astype(str).str.strip().str.lower()
        return any(z not in ("", "float", "nan", "none") for z in zones)
    except Exception:
        return False


if page == "🗺 Route Planning":
    st.header("Route Planning")
    col_up, col_mode = st.columns([3, 1])
    with col_up:
        uploaded = st.file_uploader(
            "Upload **today.xlsx** (Orders + Vehicles sheets)",
            type=["xlsx"],
            help="Need a blank template? Admin → Fleet & Zones → Download blank template.",
        )
    with col_mode:
        st.markdown("**Mode**")
        mode_label = st.radio(
            "mode", ["🧠 SMART", "🔓 FREE"], label_visibility="collapsed",
            help=(
                "**SMART** — Zone-aware. Each vehicle biased to its own customers "
                "but solver can reassign for efficiency. Daily use.\n\n"
                "**FREE** — No zone bias. Pure distance. Use as benchmark or backup."
            ),
        )
        mode = "smart" if "SMART" in mode_label else "free"

    if uploaded:
        file_bytes = uploaded.getvalue()
        try:
            prev = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Orders").dropna(how="all")
            st.caption(f"📦 **{len(prev)}** orders detected")
        except Exception:
            pass
        if mode == "smart" and not check_zones_assigned(file_bytes):
            st.warning("⚠️ No zone assignments found in Vehicles sheet. "
                       "Switch to FREE mode or set zones in the Vehicles sheet.")

        if st.button("▶  Generate Routes", type="primary"):
            max_sec   = config.SOLVER_TIME_LIMIT_SECONDS
            mode_name = "SMART" if mode == "smart" else "FREE"

            st.markdown(f"**Running {mode_name} routing...**")
            progress_bar  = st.progress(0)
            status_text   = st.empty()

            result_holder = {}
            thread = threading.Thread(
                target=run_routing_threaded,
                args=(file_bytes, mode, result_holder),
                daemon=True,
            )
            thread.start()

            start = time.time()
            phases = [
                (0.05, "Loading orders and vehicles..."),
                (0.10, "Building distance matrix..."),
                (0.18, "Finding initial solution..."),
                (0.90, f"Optimising routes (up to {max_sec}s)..."),
                (0.97, "Writing output files..."),
            ]
            phase_idx = 0

            while thread.is_alive():
                elapsed  = time.time() - start
                # Time-based progress: 0→95% over max_sec, then hold
                raw_prog = min(elapsed / max_sec, 0.95)
                # Advance through named phases
                while phase_idx < len(phases) and raw_prog >= phases[phase_idx][0]:
                    phase_idx += 1
                label = phases[min(phase_idx, len(phases)-1)][1]
                progress_bar.progress(raw_prog, text=label)
                status_text.caption(f"Elapsed: {int(elapsed)}s / {max_sec}s")
                time.sleep(0.5)

            progress_bar.progress(1.0, text="Done!")
            status_text.empty()

            if "error" in result_holder:
                st.error(f"❌ {result_holder['error']}")
            else:
                excel_bytes, routes, stops_df, unassigned, zone_summary, map_html = \
                    result_holder["result"]
                st.session_state.update({
                    "excel_bytes": excel_bytes, "routes": routes,
                    "stops_count": len(stops_df), "unassigned": unassigned,
                    "zone_summary": zone_summary, "map_html": map_html,
                })
                st.rerun()

    if "excel_bytes" in st.session_state:
        c1, c2, c3 = st.columns(3)
        c1.metric("Vehicles Routed", len(st.session_state["routes"]))
        c2.metric("Stops Assigned",  st.session_state["stops_count"])
        skipped = len(st.session_state.get("unassigned", []))
        c3.metric("Unassigned", skipped,
                  delta=f"-{skipped}" if skipped else None, delta_color="inverse")

        d1, d2 = st.columns(2)
        with d1:
            st.download_button("⬇  Download routes_output.xlsx",
                               st.session_state["excel_bytes"],
                               file_name="routes_output.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               type="primary", use_container_width=True)
        with d2:
            if st.session_state.get("map_html"):
                st.download_button("🗺  Download Route Map (HTML)",
                                   st.session_state["map_html"].encode("utf-8"),
                                   file_name="route_map.html", mime="text/html",
                                   use_container_width=True)

        if skipped:
            st.warning(f"⚠️ **{skipped}** stop(s) unassigned — see Unassigned sheet.")

        if st.session_state.get("map_html"):
            st.markdown("---")
            st.subheader("Interactive Route Map")
            st.caption("Filter by vehicle type, vehicle name, or trip number in the map sidebar.")
            components.html(st.session_state["map_html"], height=540, scrolling=False)

        zs = st.session_state.get("zone_summary")
        if zs:
            st.markdown("---")
            st.subheader("Zone Load Summary")
            df_zs = pd.DataFrame(zs)[["vehicle","zone","orig_stops","stops","kg","status"]].copy()
            df_zs.columns = ["Vehicle","Zone","Original","Final","KG","Status"]
            def _color(row):
                c = {"OVERLOADED":"background-color:#FFDDC1","LIGHT":"background-color:#FFFACD",
                     "NO ORDERS":"background-color:#E8E8E8"}.get(row["Status"],"")
                return [c]*len(row)
            st.dataframe(df_zs.style.apply(_color, axis=1),
                         use_container_width=True, hide_index=True, height=380)


elif page == "⚙️ Admin":
    st.header("Admin Panel")
    try:
        admin_pw = st.secrets["admin_password"]
    except Exception:
        admin_pw = "dsd2024"

    if not st.session_state.get("admin_auth"):
        with st.form("login_form"):
            pw = st.text_input("Password", type="password")
            if st.form_submit_button("Login", type="primary"):
                if pw == admin_pw:
                    st.session_state["admin_auth"] = True
                    st.rerun()
                else:
                    st.error("Incorrect password.")
        st.stop()

    col_s, col_l = st.columns([5, 1])
    col_s.success("✅ Logged in as Admin")
    if col_l.button("Logout"):
        st.session_state["admin_auth"] = False
        st.rerun()
    st.markdown("---")

    tab_master, tab_customer, tab_fleet, tab_zonebuilder, tab_cfg = st.tabs([
        "📋 Customer Master", "➕ New Customer", "🚛 Fleet Registry",
        "🗺 Zone Builder", "⚙️ Configuration"])

    with tab_master:
        st.subheader("Customer Master")
        mp = config.CUSTOMER_MASTER
        if os.path.exists(mp):
            try:
                mdf = pd.read_excel(mp, sheet_name="Customers")
                c1,c2,c3 = st.columns(3)
                c1.metric("Customers", len(mdf))
                c2.metric("Zones", mdf["zone"].nunique() if "zone" in mdf.columns else "—")
                visits_col = next((c for c in mdf.columns if "visit" in c.lower()), None)
                c3.metric("With History", int((mdf[visits_col]>0).sum()) if visits_col else "—")
                disp_cols = [c for c in ["customer_code","customer_name","zone","preferred_vehicle","avg_weight_kg"] if c in mdf.columns]
                st.dataframe(mdf[disp_cols].head(500), use_container_width=True, hide_index=True, height=300)
            except Exception as e:
                st.error(f"Error reading customer master: {e}")
        else:
            st.info("No customer master uploaded yet. Use the Upload button below.")
        st.markdown("---")
        cu, cd = st.columns(2)
        with cu:
            st.markdown("**Upload New Master**")
            new_m = st.file_uploader("Upload customer_master.xlsx", type=["xlsx"], key="master_up")
            if new_m and st.button("✅ Replace", type="primary"):
                os.makedirs(os.path.dirname(mp), exist_ok=True)
                with open(mp, "wb") as f: f.write(new_m.getvalue())
                st.success("Customer master replaced.")
                st.rerun()
        with cd:
            st.markdown("**Download Current Master**")
            if os.path.exists(mp):
                with open(mp, "rb") as f:
                    st.download_button("⬇ Download", f.read(), "customer_master.xlsx",
                                       use_container_width=True)

    with tab_customer:
        st.subheader("Add New Customer")
        mp2 = config.CUSTOMER_MASTER
        if not os.path.exists(mp2):
            st.info("Upload a customer master first (Customer Master tab).")
        else:
            try:
                mdf2 = pd.read_excel(mp2, sheet_name="Customers")
                mdf2["customer_code"] = mdf2["customer_code"].astype(str)
                zones_list = sorted(mdf2["zone"].dropna().unique().tolist()) if "zone" in mdf2.columns else ["Float"]
            except Exception as e:
                st.error(f"Could not read customer master: {e}")
                zones_list = ["Float"]
                mdf2 = pd.DataFrame()
            with st.form("new_cust_form"):
                col1, col2 = st.columns(2)
                with col1:
                    code=st.text_input("Customer Code *"); name=st.text_input("Customer Name *")
                    street=st.text_input("Street")
                    lat=st.number_input("Latitude *", value=42.005, format="%.6f", step=0.0001)
                    lon=st.number_input("Longitude *", value=21.435, format="%.6f", step=0.0001)
                with col2:
                    zone=st.selectbox("Zone (= vehicle name) *", zones_list)
                    pref=st.selectbox("Preferred Vehicle", ["Kamion","Furgon","Van"])
                    eligible=st.multiselect("Eligible Vehicles *", ["Kamion","Furgon","Van"], default=[pref])
                    tw_s=st.text_input("TW Start","06:00"); tw_e=st.text_input("TW End","18:00")
                if st.form_submit_button("Add Customer", type="primary"):
                    if not code or not name or not eligible:
                        st.error("Fill all required fields.")
                    elif code in mdf2["customer_code"].values:
                        st.error(f"Code {code} already exists.")
                    else:
                        wb_m=openpyxl.load_workbook(mp2); ws_m=wb_m["Customers"]
                        cols=[c.value for c in ws_m[1]]
                        nrow={"customer_code":code,"customer_name":name,"street":street,
                              "latitude":lat,"longitude":lon,"zone":zone,"special_zone":"",
                              "eligible_vehicles":",".join(eligible),"preferred_vehicle":pref,
                              "time_window_start":tw_s,"time_window_end":tw_e,
                              "visits":0,"avg_weight_kg":0,"avg_cases":0,"kg_per_case":0,
                              "vehicle_breakdown":"New customer","notes":"Added via web admin"}
                        ws_m.append([nrow.get(c,"") for c in cols])
                        wb_m.save(mp2)
                        st.success(f"✅ {name} ({code}) added to zone {zone}.")

    with tab_fleet:
        st.subheader("Fleet & Zones")
        st.markdown("#### Blank today.xlsx Template")
        st.markdown("Download, fill in your vehicles once, then use daily.")
        st.download_button("⬇  Download blank today.xlsx",
                           make_blank_template(), file_name="today.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           type="primary")
        st.markdown("---")
        st.markdown("""**Vehicles sheet columns:**

| Column | Description |
|---|---|
| `vehicle_name` | Name/plate — e.g. `DSD 1`, `F 3`, `Van 7` |
| `vehicle_type` | `Kamion`, `Furgon`, or `Van` |
| `capacity_kg` | Usable load per trip in kg |
| `zone` | Same as `vehicle_name` for dedicated zone. `Float` for overflow. |
| `available` | `TRUE` = working today |
| `max_trips_per_day` | `2` normal, `3` heavy |
""")
        if os.path.exists(config.CUSTOMER_MASTER):
            try:
                xl = pd.ExcelFile(config.CUSTOMER_MASTER)
                if "Zone Summary" in xl.sheet_names:
                    mdf3 = xl.parse("Zone Summary")
                    st.markdown("**Zone Summary from Customer Master**")
                    st.dataframe(mdf3, use_container_width=True, hide_index=True, height=260)
                elif "Customers" in xl.sheet_names:
                    mdf3 = xl.parse("Customers")
                    if "zone" in mdf3.columns:
                        zsum = mdf3.groupby("zone").size().reset_index(name="customers")
                        st.markdown("**Zone Summary (from Customers sheet)**")
                        st.dataframe(zsum, use_container_width=True, hide_index=True, height=260)
            except Exception as e:
                st.warning(f"Could not load zone summary: {e}")

    with tab_zonebuilder:
        st.subheader("Zone Builder")
        st.markdown(
            "Upload historical delivery data and your fleet is read from the Fleet Registry. "
            "The model will automatically assign each customer to a geographically compact, "
            "workload-balanced zone named after your vehicles."
        )
        st.markdown("---")

        # Fleet status
        fleet_ok = os.path.exists(FLEET_FILE)
        if fleet_ok:
            try:
                fl_df = pd.read_excel(FLEET_FILE, sheet_name="Fleet")
                fl_df = fl_df.dropna(subset=["vehicle_name"])
                fl_df = fl_df[fl_df["vehicle_name"].astype(str).str.len() > 2]
                fl_df = fl_df[~fl_df["vehicle_name"].astype(str).str.startswith("HOW")]
                k = (fl_df["vehicle_type"]=="Kamion").sum()
                f = (fl_df["vehicle_type"]=="Furgon").sum()
                v = (fl_df["vehicle_type"]=="Van").sum()
                st.success(f"✅ Fleet Registry loaded: {k} Kamion · {f} Furgon · {v} Van ({len(fl_df)} total)")
            except Exception as e:
                fleet_ok = False
                st.warning(f"Could not read fleet.xlsx: {e}")
        else:
            st.warning("No fleet.xlsx found. Go to the Fleet Registry tab to define your fleet first.")

        # History upload
        hist_file = st.file_uploader(
            "Upload historical delivery file",
            type=["xlsx"],
            help="Required columns: Delivery Date, Customer Code, Customer Name, "
                 "Latitude, Longitude, Total Weight, Vehicle Type"
        )

        master_file = st.file_uploader(
            "Upload existing customer master (optional — used to preserve non-zone data)",
            type=["xlsx"], key="zb_master"
        )

        if hist_file and fleet_ok:
            if st.button("🗺  Build Zones", type="primary"):
                with st.spinner("Clustering customers into zones... (may take 30–60s)"):
                    try:
                        hist_df   = pd.read_excel(io.BytesIO(hist_file.getvalue()))
                        master_df = None
                        if master_file:
                            try:
                                master_df = pd.read_excel(
                                    io.BytesIO(master_file.getvalue()),
                                    sheet_name="Customers"
                                )
                            except Exception:
                                pass

                        updated, zone_summary, map_html, quality = fr.build_zones(
                            hist_df, fl_df, master_df
                        )
                        st.session_state.update({
                            "zb_updated":   updated,
                            "zb_summary":   zone_summary,
                            "zb_map":       map_html,
                            "zb_quality":   quality,
                        })
                    except Exception as e:
                        st.error(f"Zone building failed: {e}")

        if "zb_updated" in st.session_state:
            updated      = st.session_state["zb_updated"]
            zone_summary = st.session_state["zb_summary"]
            map_html     = st.session_state["zb_map"]

            st.markdown("---")
            st.success(f"✅ Zones built for **{len(updated):,}** customers across **{len(zone_summary)}** zones.")

            # Quality score
            q = st.session_state.get("zb_quality", {})
            if q:
                qc1, qc2, qc3, qc4 = st.columns(4)
                qc1.metric("🏆 Composite Score", f"{q.get('composite_score',0)}/100",
                           help="Overall zone quality: balance + compactness")
                qc2.metric("⚖️ Stop Balance", f"{q.get('stop_score',0)}/100",
                           help=f"CV={q.get('stop_balance_cv',0):.3f} — lower CV = more equal daily stops")
                qc3.metric("📦 Weight Balance", f"{q.get('weight_score',0)}/100",
                           help=f"CV={q.get('weight_balance_cv',0):.3f} — lower CV = more equal daily weight")
                qc4.metric("📍 Compactness", f"{q.get('compactness_score',0)}/100",
                           help=f"Avg {q.get('avg_compactness_km',0):.1f} km intra-zone radius")

            # Download updated master
            buf = io.BytesIO()
            import openpyxl as _opx
            wb_out = _opx.Workbook()
            ws_out = wb_out.active
            ws_out.title = "Customers"
            cols_out = list(updated.columns)
            ws_out.append(cols_out)
            for _, row in updated.iterrows():
                ws_out.append([row[c] for c in cols_out])
            wb_out.save(buf)
            buf.seek(0)
            st.download_button(
                "⬇  Download updated customer_master.xlsx",
                buf.getvalue(),
                file_name="customer_master.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )

            # Validation table
            st.markdown("#### Zone Validation Summary")
            st.caption("Avg = mean day stops · P75 = busy day stops (1-in-4) · Utilisation based on P75 demand")
            st.caption("✅ OK  🟡 Heavy on busy days (>70%)  ⚠️ Overloaded on busy days (>90%)  💤 Very light")
            df_sum = pd.DataFrame(zone_summary)[[
                "zone","vehicle_type","customers",
                "exp_daily_stops","p75_daily_stops","exp_daily_kg",
                "trip_capacity_kg","utilisation_pct","flag"
            ]].copy()
            df_sum.columns = [
                "Zone","Type","Customers",
                "Avg Stops/Day","P75 Stops/Day","Avg KG/Day",
                "Trip Cap (kg)","P75 Utilisation %","Status"
            ]

            def _zb_color(row):
                if "OVER" in str(row["Status"]):
                    return ["background-color:#fdd"]*len(row)
                if "Heavy" in str(row["Status"]):
                    return ["background-color:#fff9c4"]*len(row)
                if "light" in str(row["Status"]):
                    return ["background-color:#d4eaf7"]*len(row)
                return [""]*len(row)

            st.dataframe(
                df_sum.style.apply(_zb_color, axis=1),
                use_container_width=True, hide_index=True, height=400
            )

            # Interactive map
            st.markdown("#### Interactive Zone Map")
            st.caption("Filter by vehicle type or zone in the map sidebar. Summary table is shown below the map.")
            components.html(map_html, height=600, scrolling=False)

            if st.button("🗑  Clear Results"):
                for k in ["zb_updated","zb_summary","zb_map"]:
                    st.session_state.pop(k, None)
                st.rerun()

    with tab_fleet:
        st.subheader("Fleet Registry")
        st.markdown(
            "Define your fleet here once. This is the source of truth used by the Zone Builder "
            "and for generating the `today.xlsx` template."
        )

        col_up, col_dn = st.columns(2)
        with col_up:
            st.markdown("**Upload / Replace Fleet Registry**")
            new_fl = st.file_uploader("Upload fleet.xlsx", type=["xlsx"], key="fleet_up")
            if new_fl and st.button("✅ Replace Fleet Registry", type="primary"):
                os.makedirs("_setup", exist_ok=True)
                with open(FLEET_FILE, "wb") as f:
                    f.write(new_fl.getvalue())
                st.success("Fleet registry replaced.")
                st.rerun()

        with col_dn:
            st.markdown("**Download Current Fleet Registry**")
            if os.path.exists(FLEET_FILE):
                with open(FLEET_FILE, "rb") as f:
                    st.download_button("⬇ Download fleet.xlsx", f.read(),
                                       "fleet.xlsx", use_container_width=True)

        st.markdown("---")

        # Show current fleet with inline editing
        if os.path.exists(FLEET_FILE):
            try:
                fl_edit = pd.read_excel(FLEET_FILE, sheet_name="Fleet")
                fl_edit = fl_edit.dropna(subset=["vehicle_name"])
                fl_edit = fl_edit[fl_edit["vehicle_name"].astype(str).str.len() > 2]
                fl_edit = fl_edit[~fl_edit["vehicle_name"].astype(str).str.startswith("HOW")]
                fl_show = fl_edit[["vehicle_name","vehicle_type","capacity_kg","max_trips_per_day","notes"]].copy()
                fl_show.columns = ["Vehicle Name","Type","Capacity/Trip (kg)","Max Trips/Day","Notes"]

                c1,c2,c3 = st.columns(3)
                c1.metric("Kamion", (fl_edit["vehicle_type"]=="Kamion").sum())
                c2.metric("Furgon", (fl_edit["vehicle_type"]=="Furgon").sum())
                c3.metric("Van",    (fl_edit["vehicle_type"]=="Van").sum())

                st.caption("Edit the table below then click **Save Changes**.")
                edited = st.data_editor(
                    fl_show, use_container_width=True, hide_index=True,
                    num_rows="dynamic",
                    column_config={
                        "Type": st.column_config.SelectboxColumn(
                            options=["Kamion","Furgon","Van"], required=True
                        ),
                        "Capacity/Trip (kg)": st.column_config.NumberColumn(
                            min_value=100, max_value=30_000, step=100
                        ),
                        "Max Trips/Day": st.column_config.NumberColumn(
                            min_value=1, max_value=5, step=1
                        ),
                    }
                )
                if st.button("💾 Save Changes", type="primary"):
                    import openpyxl as _opx2
                    from openpyxl.styles import PatternFill as _PF, Font as _Fn, Alignment as _Al
                    _HDR = _PF("solid", fgColor="1F4E79")
                    _TF  = {"Kamion":_PF("solid",fgColor="D6E4F7"),
                            "Furgon":_PF("solid",fgColor="D6F7E4"),
                            "Van":   _PF("solid",fgColor="FFF3CD")}
                    wb2 = _opx2.Workbook(); ws2 = wb2.active; ws2.title = "Fleet"
                    hdr = ["vehicle_name","vehicle_type","capacity_kg","max_trips_per_day","notes"]
                    ws2.append(hdr)
                    for cell in ws2[1]:
                        cell.fill = _HDR
                        cell.font = _Fn(color="FFFFFF", bold=True)
                        cell.alignment = _Al(horizontal="center")
                    edited.columns = hdr
                    for _, row in edited.iterrows():
                        vt = str(row.get("vehicle_type",""))
                        ws2.append([row[c] for c in hdr])
                        if vt in _TF:
                            for cell in ws2[ws2.max_row]: cell.fill = _TF[vt]
                    for col_l, w in zip(["A","B","C","D","E"],[22,12,14,18,30]):
                        ws2.column_dimensions[col_l].width = w
                    ws2.freeze_panes = "A2"
                    wb2.save(FLEET_FILE)
                    st.success("Fleet registry saved.")
                    st.rerun()
            except Exception as e:
                st.error(f"Could not read fleet registry: {e}")
        else:
            st.info("No fleet registry yet. Upload one above or download the blank template from below.")
            st.markdown("**Download Blank Fleet Template**")
            import openpyxl as _opx3
            _buf2 = io.BytesIO()
            _wb3  = _opx3.Workbook(); _ws3 = _wb3.active; _ws3.title = "Fleet"
            _ws3.append(["vehicle_name","vehicle_type","capacity_kg","max_trips_per_day","notes"])
            _ws3.append(["Example: DSD 1","Kamion",6000,2,""])
            _wb3.save(_buf2); _buf2.seek(0)
            st.download_button("⬇ Download blank fleet.xlsx", _buf2.getvalue(),
                               "fleet.xlsx", use_container_width=True)

        st.markdown("---")
        st.markdown("**Download Blank today.xlsx template**")
        st.download_button("⬇  Download blank today.xlsx",
                           make_blank_template(), file_name="today.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with tab_cfg:
        st.subheader("Configuration")
        st.info("Edit **config.py** and push to GitHub to apply changes.")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Depot**")
            st.code(f"Lat: {config.DEPOT_LAT}  Lon: {config.DEPOT_LON}\n"
                    f"Open: {config.DEPOT_OPEN}  Close: {config.DEPOT_CLOSE}")
            st.markdown("**Capacity Defaults**")
            for vt, cap in config.TRIP_CAPACITY.items():
                st.text(f"  {vt}: {cap:,} kg/trip")
        with c2:
            st.markdown("**Solver**")
            st.code(f"Max stops/day : {config.MAX_STOPS_PER_DAY}\n"
                    f"Solver time   : {config.SOLVER_TIME_LIMIT_SECONDS} s\n"
                    f"Zone penalty  : {config.ZONE_AFFINITY_PENALTY_KM} km\n"
                    f"Avg speed     : {config.AVERAGE_SPEED_KMH} km/h")
            st.markdown("**Special Zones**")
            for zn, zd in config.SPECIAL_ZONES.items():
                st.text(f"  {zn}: {zd['primary_vehicle']}  "
                        f"{zd['time_window_start']}–{zd['time_window_end']}")

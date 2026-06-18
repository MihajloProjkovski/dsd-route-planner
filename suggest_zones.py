#!/usr/bin/env python3
"""
suggest_zones.py
----------------
Reads the Zone Summary from _setup/customer_master.xlsx and suggests a
zone assignment for every vehicle in today.xlsx, then writes it back.

Run:  python suggest_zones.py
  or  double-click suggest_zones.bat
"""

import sys
import os
import config
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


TYPE_FILLS = {
    "Kamion": PatternFill("solid", fgColor="D6E4F7"),
    "Furgon": PatternFill("solid", fgColor="D6F7E4"),
    "Van":    PatternFill("solid", fgColor="FFF3CD"),
}
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True)
FLOAT_FILL = PatternFill("solid", fgColor="F0F0F0")


def suggest_assignments():
    master = config.CUSTOMER_MASTER
    today  = config.TODAY_FILE

    if not os.path.exists(master):
        sys.exit(f"ERROR: {master} not found. Run _setup/run_setup.bat first.")
    if not os.path.exists(today):
        sys.exit(f"ERROR: {today} not found.")

    print("=" * 55)
    print("  Zone Assignment — set zone = vehicle name for all")
    print("=" * 55)
    print("\n  Every vehicle gets its own name as its zone.\n")

    wb = openpyxl.load_workbook(today)

    if "Vehicles" not in wb.sheetnames:
        sys.exit("ERROR: today.xlsx is missing the 'Vehicles' sheet.")

    ws = wb["Vehicles"]
    header = {cell.value: cell.column for cell in ws[1] if cell.value}
    name_col = header.get("vehicle_name")
    zone_col  = header.get("zone")
    type_col  = header.get("vehicle_type")

    if not name_col or not zone_col:
        sys.exit("ERROR: Vehicles sheet must have 'vehicle_name' and 'zone' columns.")

    updated = 0
    for row in ws.iter_rows(min_row=2):
        vname_cell = row[name_col - 1]
        zone_cell  = row[zone_col  - 1]
        vname = str(vname_cell.value).strip() if vname_cell.value else ""
        if not vname or vname == "None" or vname.startswith("HOW TO"):
            continue
        zone_cell.value = vname   # zone = vehicle name
        vtype = str(row[type_col - 1].value).strip() if type_col else ""
        fill  = TYPE_FILLS.get(vtype, FLOAT_FILL)
        for cell in row:
            cell.fill = fill
        updated += 1

    try:
        wb.save(today)
    except PermissionError:
        print(f"\n  ERROR: Cannot save '{today}' - the file is open in Excel.")
        print(f"  Close today.xlsx and run suggest_zones.bat again.\n")
        sys.exit(1)

    print(f"\n  Saved to '{today}' — {updated} vehicles updated.")
    print("  Zone = vehicle name. Vehicles with no name get Float.\n")


if __name__ == "__main__":
    suggest_assignments()

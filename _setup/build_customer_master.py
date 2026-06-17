#!/usr/bin/env python3
"""
build_customer_master.py
------------------------
One-time setup script (re-run monthly to refresh from new history).

Reads Za model.xlsx and produces:
  * _setup/customer_master.xlsx  - one row per customer
  * today.xlsx (in root)         - daily input file (Orders + Vehicles sheets)

Run from _setup folder:  python build_customer_master.py
  or double-click:        _setup/run_setup.bat
"""

import sys
import os
ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, ROOT)
os.chdir(ROOT)

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from sklearn.cluster import KMeans

import config

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(color="FFFFFF", bold=True)
FLAG_FILL  = PatternFill("solid", fgColor="FFD7D7")
SPEC_FILL  = PatternFill("solid", fgColor="FFF3CD")
ZONE_FILLS = [
    PatternFill("solid", fgColor="D6E4F7"),
    PatternFill("solid", fgColor="D6F7E4"),
    PatternFill("solid", fgColor="E8D6F7"),
    PatternFill("solid", fgColor="F7D6D6"),
    PatternFill("solid", fgColor="D6F7F7"),
    PatternFill("solid", fgColor="F7F0D6"),
]


def point_in_poly(lat, lon, polygon):
    n = len(polygon)
    inside = False
    x, y = lon, lat
    j = n - 1
    for i in range(n):
        xi, yi = polygon[i][1], polygon[i][0]
        xj, yj = polygon[j][1], polygon[j][0]
        if ((yi > y) != (yj > y)) and (x < (xj - xi) * (y - yi) / (yj - yi) + xi):
            inside = not inside
        j = i
    return inside


def detect_special_zone(lat, lon):
    for zone_name, zone_def in config.SPECIAL_ZONES.items():
        if point_in_poly(lat, lon, zone_def["polygon"]):
            return zone_name
    return None


def classify_eligibility(vehicle_counts: dict, total: int) -> list:
    """
    Three-rule classification:
      Furgon dominant -> Furgon + Kamion + Van
      Kamion dominant -> Kamion + Van (+ Furgon if share >= threshold)
      Van dominant    -> Van only (+ Kamion/Furgon if share >= threshold)
    All customers with any history are classified. Zero-history customers
    handled separately in load_orders (default: Kamion, Van).
    """
    furgon_pct = vehicle_counts.get("Furgon", 0) / total * 100
    kamion_pct = vehicle_counts.get("Kamion", 0) / total * 100
    dominant   = max(vehicle_counts, key=vehicle_counts.get)

    if dominant == "Furgon":
        return ["Furgon", "Kamion", "Van"]

    if dominant == "Kamion":
        eligible = ["Kamion", "Van"]
        if furgon_pct >= config.FURGON_ELIGIBLE_FROM_KAMION_PCT:
            eligible.append("Furgon")
        return sorted(eligible)

    eligible = ["Van"]
    if kamion_pct >= config.KAMION_ELIGIBLE_FROM_VAN_PCT:
        eligible.append("Kamion")
    if furgon_pct >= config.FURGON_ELIGIBLE_FROM_VAN_PCT:
        eligible.append("Furgon")
    return sorted(eligible)


def apply_zone_restriction(eligible: list, special_zone: str) -> list:
    if not special_zone:
        return eligible
    allowed = config.SPECIAL_ZONES[special_zone]["allowed_vehicles"]
    result  = sorted(set(eligible) & set(allowed))
    return result if result else allowed


def _style_header(ws):
    for cell in ws[1]:
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws, max_w=50):
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, max_w)


def main():
    print("=" * 60)
    print("  DSD Customer Master Builder")
    print("=" * 60)

    print(f"\nLoading '{config.HISTORICAL_FILE}'...")
    df = pd.read_excel(config.HISTORICAL_FILE)
    df.columns = df.columns.str.strip()

    df["Latitude"]        = pd.to_numeric(df["Latitude"],        errors="coerce")
    df["Longitude"]       = pd.to_numeric(df["Longitude"],       errors="coerce")
    df["Total Weight"]    = pd.to_numeric(df["Total Weight"],    errors="coerce")
    df["Number of Cases"] = pd.to_numeric(df["Number of Cases"], errors="coerce")

    df = df.dropna(subset=["Latitude", "Longitude", "Customer Code"])

    # Exclude known ERP data errors
    EXCLUDED_CUSTOMERS = ["9500002182"]  # ДУКАНЏИК - ERP extraction error
    df = df[~df["Customer Code"].astype(str).isin(EXCLUDED_CUSTOMERS)]

    print(f"  Rows loaded      : {len(df):,}")
    print(f"  Unique customers : {df['Customer Code'].nunique():,}")
    print(f"  Date range       : {df['Delivery Date'].min().date()} - "
          f"{df['Delivery Date'].max().date()}")
    print(f"  Vehicle types    : {sorted(df['Vehicle Type'].unique())}")

    print("\nClassifying customers...")

    def agg_customer(grp):
        vc       = grp["Vehicle Type"].value_counts().to_dict()
        total    = len(grp)
        dominant = max(vc, key=vc.get)
        eligible = classify_eligibility(vc, total)

        wt_vals   = grp["Total Weight"].dropna()
        case_vals = grp["Number of Cases"].dropna()
        avg_kg    = wt_vals.median()   if len(wt_vals)   > 0 else 0.0
        avg_case  = case_vals.median() if len(case_vals) > 0 else 0.0
        kg_per_case = avg_kg / avg_case if avg_case > 0 else 0.0

        street_mode = grp["Street"].mode() if "Street" in grp.columns else pd.Series([])
        street = street_mode.iloc[0] if len(street_mode) > 0 else ""

        breakdown = "  ".join(
            f"{k}:{v/total*100:.1f}%" for k, v in
            sorted(vc.items(), key=lambda x: -x[1])
        )

        return pd.Series({
            "customer_name":     grp["Customer Name"].mode().iloc[0],
            "street":            street,
            "latitude":          grp["Latitude"].median(),
            "longitude":         grp["Longitude"].median(),
            "visits":            total,
            "avg_weight_kg":     round(avg_kg, 1),
            "avg_cases":         round(avg_case, 1),
            "kg_per_case":       round(kg_per_case, 3),
            "preferred_vehicle": dominant,
            "eligible_vehicles": ",".join(eligible),
            "vehicle_breakdown": breakdown,
        })

    cust = df.groupby("Customer Code").apply(
        agg_customer, include_groups=False
    ).reset_index()
    cust.rename(columns={"Customer Code": "customer_code"}, inplace=True)

    print("Detecting special zones (Plostad, Carsija)...")
    cust["special_zone"] = cust.apply(
        lambda r: detect_special_zone(r["latitude"], r["longitude"]), axis=1
    )

    cust["eligible_vehicles"] = cust.apply(
        lambda r: ",".join(
            apply_zone_restriction(r["eligible_vehicles"].split(","), r["special_zone"])
        ), axis=1
    )

    def zone_preferred(row):
        if row["special_zone"] and row["special_zone"] in config.SPECIAL_ZONES:
            return config.SPECIAL_ZONES[row["special_zone"]]["primary_vehicle"]
        return row["preferred_vehicle"]

    cust["preferred_vehicle"] = cust.apply(zone_preferred, axis=1)

    def tw_start(row):
        if row["special_zone"] and row["special_zone"] in config.SPECIAL_ZONES:
            return config.SPECIAL_ZONES[row["special_zone"]]["time_window_start"]
        return "06:00"

    def tw_end(row):
        if row["special_zone"] and row["special_zone"] in config.SPECIAL_ZONES:
            return config.SPECIAL_ZONES[row["special_zone"]]["time_window_end"]
        return "18:00"

    cust["time_window_start"] = cust.apply(tw_start, axis=1)
    cust["time_window_end"]   = cust.apply(tw_end,   axis=1)

    for sz_name in config.SPECIAL_ZONES:
        cnt = (cust["special_zone"] == sz_name).sum()
        print(f"  {sz_name:12s}: {cnt} customers")

    flags = []
    for _, row in cust.iterrows():
        sz = row["special_zone"]
        if not sz:
            continue
        allowed = config.SPECIAL_ZONES[sz]["allowed_vehicles"]
        if row["preferred_vehicle"] not in allowed:
            flags.append({
                "customer_code":       row["customer_code"],
                "customer_name":       row["customer_name"],
                "zone":                sz,
                "dominant_in_history": row["preferred_vehicle"],
                "zone_allows":         ",".join(allowed),
                "breakdown":           row["vehicle_breakdown"],
                "flag": f"Dominant vehicle ({row['preferred_vehicle']}) is NOT "
                        f"allowed in {sz}. Verify access or move to adjacent zone.",
            })

    if flags:
        print(f"\n  FLAGS: {len(flags)} customer(s) with conflicting vehicle history:")
        for f in flags:
            print(f"    {f['customer_code']}  {f['customer_name'][:30]:30s}  "
                  f"[{f['zone']}]  dominant={f['dominant_in_history']}")

    total_clusters = sum(config.N_CLUSTERS_TERRITORY_PER_TYPE.values())
    print(f"\nClustering into {total_clusters} territory zones "
          f"({config.N_CLUSTERS_TERRITORY_PER_TYPE})...")

    mask_ns = cust["special_zone"].isna()
    centres = {}

    for vtype, n_clust in config.N_CLUSTERS_TERRITORY_PER_TYPE.items():
        mask_type = mask_ns & (cust["preferred_vehicle"] == vtype)
        subset    = cust.loc[mask_type]

        if len(subset) == 0:
            continue
        if len(subset) < n_clust:
            for i, idx in enumerate(subset.index):
                cust.at[idx, "zone"] = f"{vtype}_{i+1:02d}"
            continue

        coords = subset[["latitude", "longitude"]].values
        km     = KMeans(n_clusters=n_clust, random_state=42, n_init=30)
        labels = km.fit_predict(coords)

        cust.loc[mask_type, "zone"] = [f"{vtype}_{l+1:02d}" for l in labels]

        for i, centre in enumerate(km.cluster_centers_):
            centres[f"{vtype}_{i+1:02d}"] = (round(centre[0], 5), round(centre[1], 5))

    cust.loc[~mask_ns, "zone"] = cust.loc[~mask_ns, "special_zone"]
    cust["notes"] = ""

    zone_summary = cust.groupby("zone").agg(
        customers    =("customer_code", "count"),
        dominant_veh =("preferred_vehicle", lambda x: x.value_counts().index[0]),
        avg_weight   =("avg_weight_kg", "mean"),
    ).round(1).reset_index()

    zone_summary["centre_lat"] = zone_summary["zone"].map(lambda z: centres.get(z, ("",""))[0])
    zone_summary["centre_lon"] = zone_summary["zone"].map(lambda z: centres.get(z, ("",""))[1])

    print("\n  Zone map:")
    for _, z in zone_summary.iterrows():
        print(f"    {z['zone']:<15} {z['customers']:>4} customers  "
              f"dominant={z['dominant_veh']:<8}  avg_kg={z['avg_weight']:>7.1f}")

    cols = [
        "customer_code", "customer_name", "street",
        "latitude", "longitude",
        "zone", "special_zone",
        "eligible_vehicles", "preferred_vehicle",
        "time_window_start", "time_window_end",
        "visits", "avg_weight_kg", "avg_cases", "kg_per_case",
        "vehicle_breakdown", "notes",
    ]
    cust = cust[cols]

    print(f"\nWriting '{config.CUSTOMER_MASTER}'...")
    wb = openpyxl.Workbook()

    ws_c = wb.active
    ws_c.title = "Customers"
    ws_c.append(cols)
    _style_header(ws_c)

    zone_colour = {}
    for _, row in cust.iterrows():
        ws_c.append([row[c] for c in cols])
        r = ws_c.max_row
        if row["special_zone"]:
            fill = SPEC_FILL
        elif row["zone"]:
            z = row["zone"]
            if z not in zone_colour:
                zone_colour[z] = ZONE_FILLS[len(zone_colour) % len(ZONE_FILLS)]
            fill = zone_colour[z]
        else:
            fill = None
        if fill:
            for cell in ws_c[r]:
                cell.fill = fill

    _auto_width(ws_c)
    ws_c.freeze_panes = "A2"

    ws_z = wb.create_sheet("Zone Summary")
    z_cols = ["zone", "customers", "dominant_veh", "avg_weight", "centre_lat", "centre_lon"]
    ws_z.append(z_cols)
    _style_header(ws_z)
    for _, row in zone_summary.iterrows():
        ws_z.append([row[c] for c in z_cols])
    _auto_width(ws_z)

    if flags:
        ws_f = wb.create_sheet("Flags - Review")
        flag_cols = ["customer_code", "customer_name", "zone",
                     "dominant_in_history", "zone_allows", "breakdown", "flag"]
        ws_f.append(flag_cols)
        _style_header(ws_f)
        for f in flags:
            ws_f.append([f[c] for c in flag_cols])
            for cell in ws_f[ws_f.max_row]:
                cell.fill = FLAG_FILL
        _auto_width(ws_f)

    wb.save(config.CUSTOMER_MASTER)
    print(f"  {len(cust)} customers written.")
    if flags:
        print(f"  {len(flags)} flagged customers -> 'Flags - Review' sheet.")

    today_path = config.TODAY_FILE
    print(f"\nBuilding '{config.TODAY_FILE}' (daily input: Orders + Vehicles sheets)...")

    type_fills = {
        "Kamion": PatternFill("solid", fgColor="D6E4F7"),
        "Furgon": PatternFill("solid", fgColor="D6F7E4"),
        "Van":    PatternFill("solid", fgColor="FFF3CD"),
    }

    wb2 = openpyxl.Workbook()

    ws_ord = wb2.active
    ws_ord.title = "Orders"
    ws_ord.append(["customer_code", "customer_name", "cases", "kg"])
    _style_header(ws_ord)
    for col_letter, width in zip(["A", "B", "C", "D"], [18, 35, 10, 12]):
        ws_ord.column_dimensions[col_letter].width = width
    ws_ord.freeze_panes = "A2"

    ws_veh = wb2.create_sheet("Vehicles")
    v_cols = ["vehicle_name", "vehicle_type", "zone", "available", "max_trips_per_day", "notes"]
    ws_veh.append(v_cols)
    _style_header(ws_veh)
    for v in config.FLEET:
        ws_veh.append([v["name"], v["type"], "", True, config.MAX_TRIPS_NORMAL, ""])
        fill = type_fills.get(v["type"])
        if fill:
            for cell in ws_veh[ws_veh.max_row]:
                cell.fill = fill
    ws_veh.append([])
    ws_veh.append(["HOW TO USE THIS SHEET:"])
    ws_veh.append(["zone      -> Zone from _setup/customer_master.xlsx Zone Summary "
                   "(e.g. Kamion_03, Plostad, Carsija). Use Float for overflow."])
    ws_veh.append(["available -> TRUE = available today, FALSE = maintenance/absent"])
    ws_veh.append(["max_trips -> 2 = normal day, 3 = high-demand day"])
    for col in ws_veh.columns:
        ml = max((len(str(c.value or "")) for c in col), default=0)
        ws_veh.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 80)
    ws_veh.freeze_panes = "A2"

    wb2.save(today_path)
    print(f"  Saved: {today_path}")
    print(f"  Orders sheet  : blank, ready for daily orders")
    print(f"  Vehicles sheet: {len(config.FLEET)} vehicles pre-filled")

    print("\n" + "=" * 60)
    print("  Setup complete!")
    print("=" * 60)
    print("""
Next steps:
  1. Open 'today.xlsx' in the main folder.
     -> Vehicles sheet: fill the 'zone' column for each vehicle.
        Use zone names from _setup/customer_master.xlsx Zone Summary.
        Examples: Kamion_03  Plostad  Carsija  Float
     -> Carsija zone -> Van only.
     -> Plostad zone -> Kamion.

  2. Review 'Flags - Review' in _setup/customer_master.xlsx
     for any customers whose history conflicts with zone rules.

  3. Each morning:
     -> Open today.xlsx
     -> Orders sheet  : paste today's customer codes + quantities
     -> Vehicles sheet: set available=FALSE for absent vehicles
     -> Save -> double-click run.bat -> open routes_output.xlsx
""")


if __name__ == "__main__":
    main()

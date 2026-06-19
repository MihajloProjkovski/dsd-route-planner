#!/usr/bin/env python3
"""
route_planner.py — DSD Route Planner core logic.
Called by app.py (web) and via run.bat (local).
Modes: python route_planner.py optimise | territory
"""

import os
import sys
import warnings
warnings.filterwarnings("ignore")

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf-8-sig"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import config
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from math import radians, cos, sin, asin, sqrt
from ortools.constraint_solver import routing_enums_pb2, pywrapcp


PENALTY_STOP = 1_000_000
KG_SCALE     = 10


def haversine_km(lat1, lon1, lat2, lon2):
    R = 6_371.0
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat, dlon = lat2 - lat1, lon2 - lon1
    a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
    return 2 * R * asin(sqrt(max(0.0, a)))


def hhmm_to_min(t):
    if isinstance(t, (int, float)):
        return int(t)
    h, m = map(int, str(t).strip().split(":"))
    return h * 60 + m


def min_to_hhmm(minutes):
    minutes = int(minutes)
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def build_matrices(locations):
    lats = np.array([loc[0] for loc in locations])
    lons = np.array([loc[1] for loc in locations])
    lats_r = np.radians(lats)
    lons_r = np.radians(lons)
    dlat = lats_r[:, None] - lats_r[None, :]
    dlon = lons_r[:, None] - lons_r[None, :]
    a    = np.sin(dlat / 2) ** 2 + (np.cos(lats_r[:, None]) * np.cos(lats_r[None, :]) * np.sin(dlon / 2) ** 2)
    a    = np.clip(a, 0, 1)
    km   = 2 * 6_371.0 * np.arcsin(np.sqrt(a)) * config.ROAD_FACTOR
    np.fill_diagonal(km, 0.0)
    dist_m = np.maximum(0, (km * 1_000).astype(int))
    time_m = np.maximum(0, (km / config.AVERAGE_SPEED_KMH * 60).astype(int))
    np.fill_diagonal(dist_m, 0)
    np.fill_diagonal(time_m, 0)
    return dist_m.tolist(), time_m.tolist()


def load_zone_centres(master_path):
    """Load median lat/lon per zone from customer master.
    Used as geographic anchors for zone affinity in territory mode."""
    if not os.path.exists(master_path):
        return {}
    try:
        mdf = pd.read_excel(master_path, sheet_name="Customers")
        mdf.columns = mdf.columns.str.strip().str.lower().str.replace(" ", "_")
        mdf["latitude"]  = pd.to_numeric(mdf["latitude"],  errors="coerce")
        mdf["longitude"] = pd.to_numeric(mdf["longitude"], errors="coerce")
        mdf = mdf.dropna(subset=["latitude", "longitude", "zone"])
        centres = {}
        for zone, grp in mdf.groupby("zone"):
            centres[str(zone)] = (grp["latitude"].median(), grp["longitude"].median())
        return centres
    except Exception:
        return {}


def load_orders(today_path, master_path):
    if not os.path.exists(today_path):
        sys.exit(f"\nERROR: '{today_path}' not found.\n  Fill the Orders sheet and save.")
    if not os.path.exists(master_path):
        sys.exit(f"\nERROR: '{master_path}' not found.\n  Run _setup/run_setup.bat to rebuild.")

    xl     = pd.ExcelFile(today_path)
    if "Orders" not in xl.sheet_names:
        sys.exit("ERROR: today.xlsx is missing the 'Orders' sheet.")

    orders = xl.parse("Orders")
    orders.columns = orders.columns.str.strip().str.lower().str.replace(" ", "_")
    orders = orders.dropna(how="all")
    orders["customer_code"] = orders["customer_code"].astype(str).str.strip()

    if "kg" not in orders.columns:
        sys.exit("ERROR: Orders sheet must have a 'kg' column.")

    orders["kg"]    = pd.to_numeric(orders["kg"],    errors="coerce").fillna(0)
    orders["cases"] = pd.to_numeric(orders.get("cases", pd.Series(0, index=orders.index)), errors="coerce").fillna(0)

    if "customer_name" in orders.columns:
        orders.rename(columns={"customer_name": "order_name"}, inplace=True)

    master = pd.read_excel(master_path, sheet_name="Customers")
    master.columns = master.columns.str.strip().str.lower().str.replace(" ", "_")
    master["customer_code"] = master["customer_code"].astype(str).str.strip()

    master_cols = ["customer_code", "customer_name", "street", "latitude", "longitude",
                   "zone", "special_zone", "eligible_vehicles", "preferred_vehicle",
                   "time_window_start", "time_window_end"]
    master = master[[c for c in master_cols if c in master.columns]]

    merged = orders.merge(master, on="customer_code", how="left")

    if "order_name" in merged.columns:
        merged["customer_name"] = merged["customer_name"].fillna(merged["order_name"])
    merged["customer_name"] = merged["customer_name"].fillna(merged["customer_code"])

    unknown = merged[merged["latitude"].isna()]
    stops   = merged[merged["latitude"].notna()].copy()

    if not unknown.empty:
        print(f"\n  WARNING: {len(unknown)} order(s) not in customer master -> SKIPPED:")
        for _, row in unknown.iterrows():
            name = row.get("customer_name", row.get("order_name", "?"))
            print(f"    Code: {row['customer_code']}  Name: {name}  KG: {row['kg']:.1f}")
        print("  -> Add these customers to customer_master.xlsx to include them.")

    stops = stops[stops["kg"] > 0].copy()

    if stops.empty:
        sys.exit("No valid orders after master lookup. Check customer codes.")

    stops["eligible_vehicles"] = stops["eligible_vehicles"].fillna(
        ",".join(config.NEW_CUSTOMER_DEFAULT_VEHICLES)
    )
    stops["preferred_vehicle"] = stops["preferred_vehicle"].fillna("Van")
    stops["zone"]              = stops["zone"].fillna("Unknown")
    stops["special_zone"]      = stops["special_zone"].fillna("")
    stops["time_window_start"] = stops["time_window_start"].fillna("06:00")
    stops["time_window_end"]   = stops["time_window_end"].fillna("18:00")
    stops["street"]            = stops["street"].fillna("") if "street" in stops.columns else ""

    return stops.reset_index(drop=True)


def load_vehicles(today_path):
    xl = pd.ExcelFile(today_path)
    if "Vehicles" not in xl.sheet_names:
        sys.exit("ERROR: today.xlsx is missing the 'Vehicles' sheet.")

    veh = xl.parse("Vehicles")
    # Accept both old (6-col) and new (7-col with capacity_kg) sheet layouts
    veh = veh[veh.columns[:7]]
    veh.columns = veh.columns.str.strip().str.lower().str.replace(" ", "_")
    veh = veh.dropna(subset=["vehicle_name"])
    veh["vehicle_name"] = veh["vehicle_name"].astype(str).str.strip()
    veh["vehicle_type"] = veh["vehicle_type"].astype(str).str.strip()
    veh["zone"]         = veh["zone"].fillna("Float").astype(str).str.strip()

    def parse_bool(v):
        if v is None: return False
        if isinstance(v, bool): return v
        if isinstance(v, (int, float)): return bool(v) and v == v
        return str(v).strip().lower() in ("true", "yes", "1", "y")

    veh["available"]         = veh["available"].apply(parse_bool)
    veh["max_trips_per_day"] = pd.to_numeric(veh.get("max_trips_per_day", 2), errors="coerce").fillna(config.MAX_TRIPS_NORMAL).astype(int)

    # Per-vehicle capacity: use column if present, else fall back to type default
    def resolve_capacity(row):
        cap_col = pd.to_numeric(row.get("capacity_kg", None), errors="coerce")
        if pd.notna(cap_col) and cap_col > 0:
            return int(cap_col)
        return config.TRIP_CAPACITY.get(str(row.get("vehicle_type", "Van")), 3_200)

    veh["capacity_kg"] = veh.apply(resolve_capacity, axis=1)

    available = veh[veh["available"]].reset_index(drop=True)
    if available.empty:
        sys.exit("ERROR: No vehicles are marked as available in today.xlsx Vehicles sheet.")
    return available


def solve(stops_df, vehicles_df, zone_affinity=False):
    depot     = (config.DEPOT_LAT, config.DEPOT_LON)
    stop_locs = list(zip(stops_df["latitude"], stops_df["longitude"]))
    locations = [depot] + stop_locs
    n_nodes   = len(locations)

    veh_list = []
    for _, row in vehicles_df.iterrows():
        vtype     = row["vehicle_type"]
        # Use per-vehicle capacity if set, else fall back to type default
        trip_cap  = int(row.get("capacity_kg", 0) or 0) or config.TRIP_CAPACITY.get(vtype, 3_200)
        max_trips = int(row["max_trips_per_day"])
        veh_list.append({
            "name":      row["vehicle_name"],
            "type":      vtype,
            "zone":      row["zone"],
            "trip_cap":  trip_cap,
            "daily_cap": trip_cap * max_trips,
            "max_trips": max_trips,
        })
    n_veh      = len(veh_list)
    veh_types  = [v["type"] for v in veh_list]

    available_types = set(vehicles_df["vehicle_type"].unique())
    skipped_pre = []
    keep_mask   = []
    for i, row in stops_df.iterrows():
        ev = [x.strip() for x in str(row["eligible_vehicles"]).split(",")]
        if not any(vt in available_types for vt in ev):
            skipped_pre.append(i)
            keep_mask.append(False)
        else:
            keep_mask.append(True)

    if skipped_pre:
        print(f"\n  WARNING: {len(skipped_pre)} stop(s) require unavailable vehicle types -> skipped")

    stops_df  = stops_df[keep_mask].reset_index(drop=True)
    if stops_df.empty:
        sys.exit("No stops left after filtering.")

    stop_locs = list(zip(stops_df["latitude"], stops_df["longitude"]))
    locations = [depot] + stop_locs
    n_nodes   = len(locations)

    def allowed_indices(stop_row_idx):
        ev = [x.strip() for x in str(stops_df.iloc[stop_row_idx]["eligible_vehicles"]).split(",")]
        return [vi for vi, vt in enumerate(veh_types) if vt in ev]

    dist_matrix, time_matrix = build_matrices(locations)

    demands    = [0] + [max(1, int(round(stops_df.iloc[i]["kg"] * KG_SCALE))) for i in range(len(stops_df))]
    capacities = [int(v["daily_cap"] * KG_SCALE) for v in veh_list]

    depot_open  = hhmm_to_min(config.DEPOT_OPEN)
    depot_close = hhmm_to_min(config.DEPOT_CLOSE)

    time_windows = [(depot_open, depot_close)]
    for _, row in stops_df.iterrows():
        tw_s = hhmm_to_min(row["time_window_start"])
        tw_e = hhmm_to_min(row["time_window_end"])
        tw_s = max(depot_open, min(tw_s, depot_close))
        tw_e = max(depot_open, min(tw_e, depot_close))
        if tw_s > tw_e: tw_e = tw_s
        time_windows.append((tw_s, tw_e))

    service_times = [0] + [config.AVG_SERVICE_MIN] * len(stops_df)

    manager = pywrapcp.RoutingIndexManager(n_nodes, n_veh, 0)
    routing = pywrapcp.RoutingModel(manager)

    def dist_cb(from_idx, to_idx):
        return dist_matrix[manager.IndexToNode(from_idx)][manager.IndexToNode(to_idx)]

    dist_cb_idx = routing.RegisterTransitCallback(dist_cb)

    if zone_affinity:
        # ── Per-vehicle arc costs with zone affinity penalty ──────────────────
        # Stops assigned to a vehicle outside its zone pay a distance penalty
        # equivalent to ZONE_AFFINITY_PENALTY_KM. This strongly discourages
        # cross-zone assignments but allows them when routing math demands it.
        penalty_m = int(config.ZONE_AFFINITY_PENALTY_KM * 1_000)
        stop_zones = [str(stops_df.iloc[i].get("zone", "")) for i in range(len(stops_df))]

        for v_idx, vehicle in enumerate(veh_list):
            vzone    = str(vehicle["zone"]).strip()
            is_float = vzone.lower() in ("float", "none", "", "nan")

            if is_float:
                # Float vehicles: no zone preference, use standard distance
                routing.SetArcCostEvaluatorOfVehicle(dist_cb_idx, v_idx)
            else:
                def make_zone_cb(v_zone):
                    def cb(from_idx, to_idx):
                        i = manager.IndexToNode(from_idx)
                        j = manager.IndexToNode(to_idx)
                        d = dist_matrix[i][j]
                        if 0 < j <= len(stop_zones):
                            if stop_zones[j - 1] != v_zone:
                                d += penalty_m
                        return d
                    return cb
                cb_idx = routing.RegisterTransitCallback(make_zone_cb(vzone))
                routing.SetArcCostEvaluatorOfVehicle(cb_idx, v_idx)
    else:
        routing.SetArcCostEvaluatorOfAllVehicles(dist_cb_idx)

    def time_cb(from_idx, to_idx):
        i = manager.IndexToNode(from_idx)
        j = manager.IndexToNode(to_idx)
        return time_matrix[i][j] + service_times[i]

    time_cb_idx = routing.RegisterTransitCallback(time_cb)
    routing.AddDimension(time_cb_idx, 60, 24 * 60, False, "Time")
    time_dim = routing.GetDimensionOrDie("Time")
    time_dim.SetGlobalSpanCostCoefficient(10)

    for node in range(1, n_nodes):
        idx = manager.NodeToIndex(node)
        time_dim.CumulVar(idx).SetRange(*time_windows[node])

    cp_solver = routing.solver()
    max_shift = config.MAX_DRIVER_HOURS * 60
    for v in range(n_veh):
        s_idx = routing.Start(v)
        e_idx = routing.End(v)
        trips = veh_list[v]["max_trips"]
        time_dim.CumulVar(s_idx).SetRange(depot_open, depot_close)
        time_dim.CumulVar(e_idx).SetRange(depot_open, depot_close + max_shift * trips)
        cp_solver.Add(time_dim.CumulVar(e_idx) - time_dim.CumulVar(s_idx) <= max_shift * trips)
        routing.AddVariableMinimizedByFinalizer(time_dim.CumulVar(s_idx))
        routing.AddVariableMinimizedByFinalizer(time_dim.CumulVar(e_idx))

    def demand_cb(from_idx):
        return demands[manager.IndexToNode(from_idx)]

    demand_cb_idx = routing.RegisterUnaryTransitCallback(demand_cb)
    routing.AddDimensionWithVehicleCapacity(demand_cb_idx, 0, capacities, True, "Capacity")

    def stop_count_cb(from_idx):
        return 0 if manager.IndexToNode(from_idx) == 0 else 1

    stop_cb_idx = routing.RegisterUnaryTransitCallback(stop_count_cb)
    routing.AddDimensionWithVehicleCapacity(
        stop_cb_idx, 0, [config.MAX_STOPS_PER_DAY] * n_veh, True, "StopCount"
    )

    for node in range(1, n_nodes):
        allowed = allowed_indices(node - 1)
        if allowed and len(allowed) < n_veh:
            routing.VehicleVar(manager.NodeToIndex(node)).SetValues([-1] + allowed)

    for node in range(1, n_nodes):
        routing.AddDisjunction([manager.NodeToIndex(node)], PENALTY_STOP)

    params = pywrapcp.DefaultRoutingSearchParameters()
    params.first_solution_strategy    = routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
    params.local_search_metaheuristic = routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    params.time_limit.seconds         = config.SOLVER_TIME_LIMIT_SECONDS

    solution = routing.SolveWithParameters(params)

    if solution is None:
        sys.exit("Solver returned no feasible solution.")

    routes            = []
    unassigned_indices = []

    for v_idx, vehicle in enumerate(veh_list):
        idx            = routing.Start(v_idx)
        stops_in_route = []

        while not routing.IsEnd(idx):
            node = manager.IndexToNode(idx)
            if node != 0:
                arrival = solution.Min(time_dim.CumulVar(idx))
                row     = stops_df.iloc[node - 1]
                stops_in_route.append({
                    "stop_order":    len(stops_in_route) + 1,
                    "orig_idx":      node - 1,
                    "customer_code": row["customer_code"],
                    "customer_name": row.get("customer_name", row["customer_code"]),
                    "street":        row.get("street", ""),
                    "latitude":      row["latitude"],
                    "longitude":     row["longitude"],
                    "kg":            row["kg"],
                    "cases":         row.get("cases", 0),
                    "arrival_time":  min_to_hhmm(arrival),
                    "time_window":   f"{row['time_window_start']} - {row['time_window_end']}",
                    "zone":          row.get("zone", ""),
                    "special_zone":  row.get("special_zone", ""),
                    "eligible_veh":  row.get("eligible_vehicles", ""),
                })
            idx = solution.Value(routing.NextVar(idx))

        if stops_in_route:
            routes.append({
                "vehicle_name": vehicle["name"],
                "vehicle_type": vehicle["type"],
                "vehicle_zone": vehicle["zone"],
                "trip_cap":     vehicle["trip_cap"],
                "max_trips":    vehicle["max_trips"],
                "stops":        stops_in_route,
            })

    for node in range(1, n_nodes):
        ri = manager.NodeToIndex(node)
        if solution.Value(routing.NextVar(ri)) == ri:
            unassigned_indices.append(node - 1)

    return routes, unassigned_indices, stops_df


def split_into_trips(route):
    cap     = route["trip_cap"]
    stops   = route["stops"]
    trips   = []
    current = []
    cur_kg  = 0.0

    for stop in stops:
        kg = stop["kg"]
        if cur_kg + kg > cap and current:
            trips.append({"trip_num": len(trips) + 1, "stops": current, "trip_kg": cur_kg})
            current = [stop]
            cur_kg  = kg
        else:
            current.append(stop)
            cur_kg += kg

    if current:
        trips.append({"trip_num": len(trips) + 1, "stops": current, "trip_kg": cur_kg})

    return trips


# ── Territory helpers ─────────────────────────────────────────────────────────

def nearest_neighbor_sequence(stops_records, depot):
    remaining    = list(stops_records)
    ordered      = []
    current_pos  = depot
    current_time = hhmm_to_min(config.DEPOT_OPEN) + 30

    while remaining:
        nearest_idx = min(
            range(len(remaining)),
            key=lambda i: haversine_km(
                current_pos[0], current_pos[1],
                remaining[i]["latitude"], remaining[i]["longitude"]
            )
        )
        rec         = remaining.pop(nearest_idx)
        travel_km   = (haversine_km(current_pos[0], current_pos[1],
                                    rec["latitude"], rec["longitude"])
                       * config.ROAD_FACTOR)
        current_time += max(1, int(travel_km / config.AVERAGE_SPEED_KMH * 60))

        ordered.append({
            "stop_order":    len(ordered) + 1,
            "orig_idx":      rec.get("_orig_idx", 0),
            "customer_code": rec["customer_code"],
            "customer_name": rec.get("customer_name", rec["customer_code"]),
            "street":        rec.get("street", ""),
            "latitude":      rec["latitude"],
            "longitude":     rec["longitude"],
            "kg":            rec["kg"],
            "cases":         rec.get("cases", 0),
            "arrival_time":  min_to_hhmm(current_time),
            "time_window":   f"{rec.get('time_window_start','06:00')} - {rec.get('time_window_end','18:00')}",
            "zone":          rec.get("zone", ""),
            "special_zone":  rec.get("special_zone", ""),
            "eligible_veh":  rec.get("eligible_vehicles", ""),
        })
        current_time += config.AVG_SERVICE_MIN
        current_pos   = (rec["latitude"], rec["longitude"])

    return ordered


def solve_territory(stops_df, vehicles_df):
    max_stops = config.MAX_STOPS_PER_DAY
    min_stops = 4
    depot     = (config.DEPOT_LAT, config.DEPOT_LON)

    # Dedicated zone vehicles
    zone_to_veh = {}
    # Float vehicles (no fixed zone — serve as overflow / unzoned customer pool)
    float_vehs  = {}   # vehicle_name -> vehicle dict

    for _, veh in vehicles_df.iterrows():
        z_raw = str(veh["zone"]).strip()
        vname = veh["vehicle_name"]
        if not z_raw or z_raw.lower() in ("", "nan", "float", "none"):
            # Explicit Float — no dedicated zone, serves as overflow
            float_vehs[vname] = veh.to_dict()
            continue
        # Any other zone value (including zone = vehicle name) is a real zone
        for z in [x.strip() for x in z_raw.split(",") if x.strip()]:
            if z.lower() not in ("float", "none", "nan"):
                zone_to_veh[z] = veh.to_dict()

    # Stop lists for Float vehicles (start empty, filled by unzoned distribution)
    float_stops = {vname: [] for vname in float_vehs}

    stops_records = stops_df.to_dict("records")
    for i, rec in enumerate(stops_records):
        rec["_orig_idx"] = i

    zone_stops   = {z: [] for z in zone_to_veh}   # pre-initialize all zones
    unzoned_recs = []
    for rec in stops_records:
        z = str(rec.get("zone", "")).strip()
        if z in zone_to_veh:
            zone_stops[z].append(rec)
        else:
            unzoned_recs.append(rec)

    original_counts = {z: len(zone_stops.get(z, [])) for z in zone_to_veh}
    rebalance_notes = {}

    def zone_centre(z):
        stops = zone_stops.get(z, [])
        if not stops:
            return (config.DEPOT_LAT, config.DEPOT_LON)
        return (
            sum(s["latitude"]  for s in stops) / len(stops),
            sum(s["longitude"] for s in stops) / len(stops),
        )

    def veh_type(z):
        return zone_to_veh[z]["vehicle_type"] if z in zone_to_veh else None

    for z in zone_stops:
        zone_stops[z].sort(
            key=lambda s: haversine_km(depot[0], depot[1], s["latitude"], s["longitude"])
        )

    for _pass in range(10):
        changed = False

        overloaded  = sorted(
            [z for z in zone_to_veh if len(zone_stops.get(z, [])) > max_stops],
            key=lambda z: -len(zone_stops.get(z, []))
        )
        underloaded = sorted(
            [z for z in zone_to_veh if len(zone_stops.get(z, [])) < min_stops],
            key=lambda z: len(zone_stops.get(z, []))
        )

        if not overloaded and not underloaded:
            break

        for over_z in overloaded:
            while len(zone_stops.get(over_z, [])) > max_stops:
                ctr = zone_centre(over_z)
                donate = max(
                    zone_stops[over_z],
                    key=lambda s: haversine_km(ctr[0], ctr[1], s["latitude"], s["longitude"])
                )
                eligible = [x.strip() for x in str(donate.get("eligible_vehicles", "Van")).split(",")]
                best_z, best_dist = None, float("inf")
                for z2, veh2 in zone_to_veh.items():
                    if z2 == over_z: continue
                    if veh2["vehicle_type"] not in eligible: continue
                    if len(zone_stops.get(z2, [])) >= max_stops: continue
                    ctr2 = zone_centre(z2)
                    d = haversine_km(donate["latitude"], donate["longitude"], ctr2[0], ctr2[1])
                    if d < best_dist:
                        best_dist, best_z = d, z2
                if best_z:
                    zone_stops[over_z].remove(donate)
                    zone_stops.setdefault(best_z, []).append(donate)
                    rebalance_notes.setdefault(over_z, []).append(f"1 stop -> {best_z} (overflow relief)")
                    rebalance_notes.setdefault(best_z, []).append(f"received 1 stop from {over_z}")
                    changed = True
                else:
                    zone_stops[over_z].remove(donate)
                    unzoned_recs.append(donate)
                    changed = True

        overloaded = [z for z in zone_to_veh if len(zone_stops.get(z, [])) > max_stops]
        for under_z in underloaded:
            cur = len(zone_stops.get(under_z, []))
            capacity_left = max_stops - cur
            if capacity_left <= 0: continue
            vtype = veh_type(under_z)
            ctr_u = zone_centre(under_z)
            candidates = [
                (z, haversine_km(ctr_u[0], ctr_u[1], *zone_centre(z)))
                for z in zone_to_veh
                if z != under_z and len(zone_stops.get(z, [])) > max_stops
                and zone_to_veh[z]["vehicle_type"] == vtype
            ]
            if not candidates:
                candidates = [
                    (z, haversine_km(ctr_u[0], ctr_u[1], *zone_centre(z)))
                    for z in zone_to_veh
                    if z != under_z and len(zone_stops.get(z, [])) > 8
                    and zone_to_veh[z]["vehicle_type"] == vtype
                ]
            if not candidates: continue
            best_donor = min(candidates, key=lambda x: x[1])[0]
            donor_can_spare = max(0, len(zone_stops.get(best_donor, [])) - min_stops)
            take_count      = min(capacity_left, donor_can_spare)
            available = [
                s for s in zone_stops.get(best_donor, [])
                if vtype in [x.strip() for x in str(s.get("eligible_vehicles", "Van")).split(",")]
            ]
            transfer = sorted(
                available,
                key=lambda s: haversine_km(ctr_u[0], ctr_u[1], s["latitude"], s["longitude"])
            )[:take_count]
            for s in transfer:
                zone_stops[best_donor].remove(s)
                zone_stops.setdefault(under_z, []).append(s)
                changed = True
            if transfer:
                n = len(transfer)
                rebalance_notes.setdefault(under_z, []).append(f"received {n} stop(s) from {best_donor} (rebalance)")
                rebalance_notes.setdefault(best_donor, []).append(f"{n} stop(s) -> {under_z} (rebalance)")

        if not changed:
            break

    # ── Distribute unzoned stops (new/unknown-zone customers) ─────────────────
    # Priority order: 1) Float vehicles (no dedicated zone)
    #                 2) Zone vehicles that are significantly underutilized
    if unzoned_recs:
        def stop_centre(stops_list):
            if not stops_list:
                return (config.DEPOT_LAT, config.DEPOT_LON)
            return (sum(s["latitude"] for s in stops_list) / len(stops_list),
                    sum(s["longitude"] for s in stops_list) / len(stops_list))

        for rec in list(unzoned_recs):
            eligible = [x.strip() for x in str(rec.get("eligible_vehicles", "Van")).split(",")]
            best_key  = None
            best_score = float("inf")
            is_float_best = False

            # Option 1 — Float vehicles (full capacity available, top priority)
            for vname, veh in float_vehs.items():
                if veh["vehicle_type"] not in eligible:
                    continue
                cur = len(float_stops.get(vname, []))
                if cur >= max_stops:
                    continue
                ctr  = stop_centre(float_stops.get(vname, []))
                dist = haversine_km(rec["latitude"], rec["longitude"], ctr[0], ctr[1])
                if dist < best_score:
                    best_score    = dist
                    best_key      = vname
                    is_float_best = True

            # Option 2 — Underutilized zone vehicles (< half of max_stops used)
            for z, veh in zone_to_veh.items():
                if veh["vehicle_type"] not in eligible:
                    continue
                cur  = len(zone_stops.get(z, []))
                orig = original_counts.get(z, cur)
                if cur >= max_stops:
                    continue
                if orig > max_stops // 2:
                    continue  # Only significantly underutilized zone vehicles
                ctr  = stop_centre(zone_stops.get(z, []))
                dist = haversine_km(rec["latitude"], rec["longitude"], ctr[0], ctr[1])
                # 100 km equivalent penalty so Float vehicles are strongly preferred
                score = 100 + dist
                if score < best_score:
                    best_score    = score
                    best_key      = z
                    is_float_best = False

            if best_key is not None:
                if is_float_best:
                    float_stops[best_key].append(rec)
                    rebalance_notes.setdefault(f"__float_{best_key}", []).append(
                        f"unzoned customer {rec.get('customer_code','?')} auto-assigned")
                else:
                    zone_stops[best_key].append(rec)
                    rebalance_notes.setdefault(best_key, []).append(
                        f"unzoned customer {rec.get('customer_code','?')} auto-assigned")
                unzoned_recs.remove(rec)

    zone_summary = []
    for z, veh in sorted(zone_to_veh.items()):
        stops_in   = zone_stops.get(z, [])
        orig_count = original_counts.get(z, 0)
        total_kg   = sum(s["kg"] for s in stops_in)
        n_stops    = len(stops_in)

        if orig_count == 0 and n_stops == 0:
            status = "NO ORDERS"
        elif n_stops > max_stops:
            status = "OVERLOADED"
        elif n_stops <= 2:
            status = "LIGHT"
        else:
            status = "OK"

        notes = rebalance_notes.get(z, [])
        zone_summary.append({
            "vehicle":    veh["vehicle_name"],
            "type":       veh["vehicle_type"],
            "zone":       z,
            "orig_stops": orig_count,
            "stops":      n_stops,
            "kg":         round(total_kg, 1),
            "status":     status,
            "note":       "  |  ".join(notes),
        })

    # Add Float vehicle entries to zone_summary (only those that received stops)
    for vname, veh in float_vehs.items():
        fl_stops = float_stops.get(vname, [])
        if not fl_stops:
            continue
        notes = rebalance_notes.get(f"__float_{vname}", [])
        zone_summary.append({
            "vehicle":    vname,
            "type":       veh["vehicle_type"],
            "zone":       "Float",
            "orig_stops": 0,
            "stops":      len(fl_stops),
            "kg":         round(sum(s["kg"] for s in fl_stops), 1),
            "status":     "OK",
            "note":       "  |  ".join(notes),
        })

    routes = []
    for z, veh in zone_to_veh.items():
        stops_list = zone_stops.get(z, [])
        if not stops_list:
            continue
        ordered  = nearest_neighbor_sequence(stops_list, depot)
        trip_cap = int(veh.get("capacity_kg", 0) or 0) or config.TRIP_CAPACITY.get(veh["vehicle_type"], 3_200)
        routes.append({
            "vehicle_name": veh["vehicle_name"],
            "vehicle_type": veh["vehicle_type"],
            "vehicle_zone": z,
            "trip_cap":     trip_cap,
            "max_trips":    int(veh["max_trips_per_day"]),
            "stops":        ordered,
        })

    # Float vehicle routes (vehicles that received unzoned / overflow stops)
    for vname, veh in float_vehs.items():
        fl_stops = float_stops.get(vname, [])
        if not fl_stops:
            continue
        ordered  = nearest_neighbor_sequence(fl_stops, depot)
        trip_cap = int(veh.get("capacity_kg", 0) or 0) or config.TRIP_CAPACITY.get(veh["vehicle_type"], 3_200)
        routes.append({
            "vehicle_name": vname,
            "vehicle_type": veh["vehicle_type"],
            "vehicle_zone": "Float",
            "trip_cap":     trip_cap,
            "max_trips":    int(veh["max_trips_per_day"]),
            "stops":        ordered,
        })

    unassigned_indices = [r["_orig_idx"] for r in unzoned_recs]
    return routes, unassigned_indices, zone_summary


# ── Excel output ──────────────────────────────────────────────────────────────

HDR_FILL      = PatternFill("solid", fgColor="1F4E79")
HDR_FONT      = Font(color="FFFFFF", bold=True)
TYPE_FILLS    = {
    "Kamion": PatternFill("solid", fgColor="D6E4F7"),
    "Furgon": PatternFill("solid", fgColor="D6F7E4"),
    "Van":    PatternFill("solid", fgColor="FFF3CD"),
}
SPECIAL_FILL  = PatternFill("solid", fgColor="FFD7D7")
TRIP_HDR_FILL = PatternFill("solid", fgColor="2E75B6")
OVERLOAD_FILL = PatternFill("solid", fgColor="FFDDC1")
LIGHT_FILL    = PatternFill("solid", fgColor="FFFACD")
NO_ORD_FILL   = PatternFill("solid", fgColor="E8E8E8")


# ── Route map generator ───────────────────────────────────────────────────────

def generate_route_map(routes) -> str:
    """Generate a self-contained Leaflet HTML map for the solved routes.
    Returns HTML string. Filters: vehicle name, vehicle type, trip number."""
    import json

    # Assign a colour per vehicle (cycle through palette)
    PALETTE = [
        "#1A5276","#1E8449","#D35400","#8E44AD","#C0392B","#17A589",
        "#B7950B","#2980B9","#27AE60","#E67E22","#6C3483","#E74C3C",
        "#1ABC9C","#F39C12","#5D6D7E","#A93226","#1F618D","#2E86C1",
        "#52BE80","#CA6F1E","#7D3C98","#117A65","#BA4A00","#154360",
    ]
    veh_colours = {}
    colour_idx  = 0
    for route in routes:
        vname = route["vehicle_name"]
        if vname not in veh_colours:
            veh_colours[vname] = PALETTE[colour_idx % len(PALETTE)]
            colour_idx += 1

    # Build GeoJSON features: one per stop
    features = []
    all_vehicles   = sorted({r["vehicle_name"] for r in routes})
    all_types      = sorted({r["vehicle_type"]  for r in routes})
    all_trip_nums  = sorted({t["trip_num"]
                             for r in routes
                             for t in split_into_trips(r)})

    for route in routes:
        trips = split_into_trips(route)
        for trip in trips:
            for stop in trip["stops"]:
                features.append({
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [stop["longitude"], stop["latitude"]],
                    },
                    "properties": {
                        "vehicle":  route["vehicle_name"],
                        "vtype":    route["vehicle_type"],
                        "trip":     trip["trip_num"],
                        "stop":     stop["stop_order"],
                        "name":     stop["customer_name"],
                        "street":   stop.get("street", ""),
                        "kg":       round(stop["kg"], 1),
                        "cases":    int(stop.get("cases", 0)),
                        "arrival":  stop["arrival_time"],
                        "colour":   veh_colours[route["vehicle_name"]],
                    }
                })

    geojson_js  = json.dumps({"type": "FeatureCollection", "features": features})
    vehicles_js = json.dumps(all_vehicles)
    types_js    = json.dumps(all_types)
    trips_js    = json.dumps(all_trip_nums)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>DSD Route Map</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',sans-serif;display:flex;height:100vh;overflow:hidden}}
  #sidebar{{width:260px;background:#1F2D3D;color:#ECF0F1;display:flex;flex-direction:column;overflow:hidden;box-shadow:2px 0 8px rgba(0,0,0,.4)}}
  #sidebar h1{{font-size:14px;font-weight:700;padding:12px 14px 8px;border-bottom:1px solid #2C3E50;color:#fff}}
  #sidebar h2{{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:1px;color:#95A5A6;padding:8px 14px 3px}}
  #filter-wrap{{flex:1;overflow-y:auto;padding-bottom:10px}}
  .fs{{padding:0 14px}}.check-row{{display:flex;align-items:center;gap:7px;padding:3px 0;font-size:12px;cursor:pointer;user-select:none}}
  .check-row input{{cursor:pointer;accent-color:#3498DB;width:13px;height:13px}}
  .swatch{{width:11px;height:11px;border-radius:50%;flex-shrink:0}}
  .btn-row{{display:flex;gap:5px;padding:4px 14px}}
  .btn{{flex:1;padding:4px 0;font-size:10px;font-weight:600;border:none;border-radius:4px;cursor:pointer;background:#2C3E50;color:#BDC3C7}}
  .btn:hover{{background:#3498DB;color:#fff}}
  #stats{{padding:6px 14px;font-size:11px;color:#7F8C8D;border-top:1px solid #2C3E50}}
  #map{{flex:1}}
  .lf-popup{{font-size:12px;line-height:1.5;min-width:190px}}
</style>
</head>
<body>
<div id="sidebar">
  <h1>🚛 DSD Route Map</h1>
  <div id="filter-wrap">
    <h2>Vehicle Type</h2>
    <div class="fs" id="type-f"></div>
    <div class="btn-row">
      <button class="btn" onclick="tAll('type',true)">All</button>
      <button class="btn" onclick="tAll('type',false)">None</button>
    </div>
    <h2>Vehicle</h2>
    <div class="fs" id="veh-f"></div>
    <div class="btn-row">
      <button class="btn" onclick="tAll('veh',true)">All</button>
      <button class="btn" onclick="tAll('veh',false)">None</button>
    </div>
    <h2>Trip #</h2>
    <div class="fs" id="trip-f"></div>
    <div class="btn-row">
      <button class="btn" onclick="tAll('trip',true)">All</button>
      <button class="btn" onclick="tAll('trip',false)">None</button>
    </div>
  </div>
  <div id="stats">Showing <b id="vc">–</b> of <b id="tc">–</b> stops</div>
</div>
<div id="map"></div>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script>
const GJ       = {geojson_js};
const VEHICLES = {vehicles_js};
const TYPES    = {types_js};
const TRIPS    = {trips_js};

const map = L.map('map').setView([41.998,21.435],13);
L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png',{{attribution:'© OpenStreetMap',maxZoom:19}}).addTo(map);

const markers=[];
GJ.features.forEach(f=>{{
  const p=f.properties;
  const mk=L.circleMarker([f.geometry.coordinates[1],f.geometry.coordinates[0]],{{
    radius:7,fillColor:p.colour,color:'#fff',weight:1.5,fillOpacity:0.9
  }});
  mk.bindPopup(`<div class="lf-popup"><b>Stop ${{p.stop}} — ${{p.vehicle}}</b><br>
    Trip ${{p.trip}} &nbsp;|&nbsp; ${{p.vtype}}<br>
    <b>${{p.name}}</b><br>
    ${{p.street?'📍 '+p.street+'<br>':''}}
    ⚖️ ${{p.kg}} kg &nbsp;|&nbsp; 📦 ${{p.cases}} cases<br>
    🕐 Arrival: ${{p.arrival}}</div>`);
  mk._v=p.vehicle; mk._t=p.vtype; mk._trip=p.trip;
  markers.push(mk);
}});
const mg=L.layerGroup(markers).addTo(map);

const aV=new Set(VEHICLES), aT=new Set(TYPES), aTr=new Set(TRIPS);
function apply(){{
  let vis=0;
  markers.forEach(m=>{{
    const ok=aT.has(m._t)&&aV.has(m._v)&&aTr.has(m._trip);
    if(ok){{if(!map.hasLayer(m))mg.addLayer(m);vis++;}}
    else{{if(map.hasLayer(m))mg.removeLayer(m);}}
  }});
  document.getElementById('vc').textContent=vis;
}}

function mkCheck(wrap,label,val,colour,group,set){{
  const lb=document.createElement('label');lb.className='check-row';
  lb.innerHTML=`<input type="checkbox" data-g="${{group}}" data-v="${{val}}" checked>
    <span class="swatch" style="background:${{colour}}"></span><span>${{label}}</span>`;
  lb.querySelector('input').addEventListener('change',e=>{{
    e.target.checked?set.add(val):set.delete(val);apply();
  }});
  document.getElementById(wrap).appendChild(lb);
}}

TYPES.forEach(t=>mkCheck('type-f',t,t,{{Kamion:'#2E86C1',Furgon:'#1E8449',Van:'#D35400'}}[t]||'#7F8C8D','type',aT));
const vclr={{}};
GJ.features.forEach(f=>{{vclr[f.properties.vehicle]=f.properties.colour;}});
VEHICLES.forEach(v=>mkCheck('veh-f',v,v,vclr[v]||'#7F8C8D','veh',aV));
TRIPS.forEach(tr=>mkCheck('trip-f','Trip '+tr,tr,'#5D6D7E','trip',aTr));

function tAll(g,s){{
  document.querySelectorAll(`input[data-g="${{g}}"]`).forEach(cb=>{{
    cb.checked=s;
    const rawVal=cb.dataset.v;
    const val=(g==='trip')?parseInt(rawVal):rawVal;
    if(g==='type')s?aT.add(val):aT.delete(val);
    else if(g==='veh')s?aV.add(val):aV.delete(val);
    else s?aTr.add(val):aTr.delete(val);
  }});apply();
}}

document.getElementById('tc').textContent=markers.length;
apply();
</script>
</body>
</html>"""
    return html


def _style_header(ws):
    for cell in ws[1]:
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws, max_w=45):
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, max_w)


def write_excel(routes, stops_df, unassigned_indices, path, zone_summary=None):
    wb = openpyxl.Workbook()

    BLOCK_FILLS = [
        PatternFill("solid", fgColor="EBF3FB"),
        PatternFill("solid", fgColor="EBF7EE"),
        PatternFill("solid", fgColor="FEF9E7"),
        PatternFill("solid", fgColor="F5EEF8"),
        PatternFill("solid", fgColor="FDFEFE"),
        PatternFill("solid", fgColor="FEF5E7"),
    ]
    VEH_HDR_FILLS = {
        "Kamion": PatternFill("solid", fgColor="2E86C1"),
        "Furgon": PatternFill("solid", fgColor="1E8449"),
        "Van":    PatternFill("solid", fgColor="D35400"),
    }

    # Sheet 1: Plan Overview
    ws_ov = wb.active
    ws_ov.title = "Plan Overview"

    if zone_summary:
        ws_ov.append(["Vehicle", "Type", "Zone",
                      "Original Stops", "Final Stops", "KG Today",
                      "Status", "Notes"])
        _style_header(ws_ov)
        status_fills = {
            "OVERLOADED": OVERLOAD_FILL,
            "LIGHT":      LIGHT_FILL,
            "NO ORDERS":  NO_ORD_FILL,
        }
        for zrow in zone_summary:
            ws_ov.append([
                zrow["vehicle"], zrow["type"], zrow["zone"],
                zrow.get("orig_stops", zrow["stops"]),
                zrow["stops"], zrow["kg"],
                zrow["status"], zrow["note"],
            ])
            fill = status_fills.get(zrow["status"], TYPE_FILLS.get(zrow["type"]))
            if fill:
                for cell in ws_ov[ws_ov.max_row]:
                    cell.fill = fill
    else:
        ws_ov.append(["Vehicle", "Type", "Zone", "Trips",
                      "Total KG", "Cap/Trip (kg)", "Utilisation %", "Stops"])
        _style_header(ws_ov)
        for route in routes:
            trips      = split_into_trips(route)
            total_kg   = sum(s["kg"] for s in route["stops"])
            trips_used = len(trips)
            util       = total_kg / (route["trip_cap"] * trips_used) * 100 if trips_used else 0
            fill       = TYPE_FILLS.get(route["vehicle_type"])
            ws_ov.append([
                route["vehicle_name"], route["vehicle_type"], route["vehicle_zone"],
                trips_used, round(total_kg, 1), route["trip_cap"],
                f"{util:.1f}%", len(route["stops"]),
            ])
            if fill:
                for cell in ws_ov[ws_ov.max_row]:
                    cell.fill = fill

    _auto_width(ws_ov)
    ws_ov.freeze_panes = "A2"

    # Sheet 2: Route Plan
    ws_rp = wb.create_sheet("Route Plan")
    ws_rp.append(["Vehicle", "Trip", "Stop #",
                  "Customer Name", "Street",
                  "KG", "Cases", "Arrival", "Time Window"])
    _style_header(ws_rp)

    col_widths = [22, 6, 6, 35, 35, 9, 7, 9, 16]
    for i, w in enumerate(col_widths, 1):
        ws_rp.column_dimensions[get_column_letter(i)].width = w

    block_idx = 0
    for route in sorted(routes, key=lambda r: r["vehicle_name"]):
        trips      = split_into_trips(route)
        total_kg   = sum(s["kg"] for s in route["stops"])
        total_stop = len(route["stops"])
        vtype      = route["vehicle_type"]
        vname      = route["vehicle_name"]
        block_fill = BLOCK_FILLS[block_idx % len(BLOCK_FILLS)]
        veh_hdr_fill = VEH_HDR_FILLS.get(vtype, HDR_FILL)
        block_idx += 1

        ws_rp.append([
            f"  {vname}  -  {route['vehicle_zone']}  "
            f"({total_stop} stops  |  {total_kg:,.0f} kg)",
            "", "", "", "", "", "", "", "",
        ])
        for cell in ws_rp[ws_rp.max_row]:
            cell.fill = veh_hdr_fill
            cell.font = Font(color="FFFFFF", bold=True, size=11)
        ws_rp.row_dimensions[ws_rp.max_row].height = 18

        for trip in trips:
            ws_rp.append([
                "",
                f"Trip {trip['trip_num']}",
                "",
                f"{trip['trip_kg']:,.0f} kg  /  {route['trip_cap']:,} kg capacity",
                "", "", "", "", "",
            ])
            for cell in ws_rp[ws_rp.max_row]:
                cell.fill = TRIP_HDR_FILL
                cell.font = Font(color="FFFFFF", bold=True)

            for i, stop in enumerate(trip["stops"]):
                fill = SPECIAL_FILL if stop.get("special_zone") else block_fill
                ws_rp.append([
                    vname, trip["trip_num"], i + 1,
                    stop["customer_name"], stop.get("street", ""),
                    round(stop["kg"], 1), int(stop.get("cases", 0)),
                    stop["arrival_time"], stop["time_window"],
                ])
                for cell in ws_rp[ws_rp.max_row]:
                    cell.fill = fill
                    cell.font = Font(size=10)

        ws_rp.append([""] * 9)

    ws_rp.freeze_panes = "A2"

    # Sheet 3: Unassigned
    if unassigned_indices:
        ws_ua = wb.create_sheet("Unassigned")
        ws_ua.append(["Customer Code", "Customer Name", "KG",
                      "Zone", "Eligible Vehicles", "Reason"])
        _style_header(ws_ua)
        for idx in unassigned_indices:
            if idx >= len(stops_df): continue
            row = stops_df.iloc[idx]
            ws_ua.append([
                row["customer_code"], row.get("customer_name", ""),
                round(row["kg"], 1), row.get("zone", ""),
                row.get("eligible_vehicles", ""),
                "No eligible vehicle could serve this stop",
            ])
            for cell in ws_ua[ws_ua.max_row]:
                cell.fill = PatternFill("solid", fgColor="FFD7D7")
        _auto_width(ws_ua)

    if hasattr(path, "write"):
        wb.save(path)
    else:
        try:
            wb.save(path)
        except PermissionError:
            sys.exit(
                f"\n  ERROR: Cannot save '{path}' - the file is open in Excel.\n"
                f"  Close routes_output.xlsx and run run.bat again.\n"
            )


# ── Main functions ─────────────────────────────────────────────────────────────

def _main_optimise():
    print("=" * 60)
    print("  DSD Route Planner - OPTIMISE MODE")
    print("=" * 60)

    print(f"\nLoading orders and vehicles from '{config.TODAY_FILE}'...")
    stops_df    = load_orders(config.TODAY_FILE, config.CUSTOMER_MASTER)
    vehicles_df = load_vehicles(config.TODAY_FILE)

    total_kg        = stops_df["kg"].sum()
    fleet_daily_cap = sum(
        config.TRIP_CAPACITY.get(row["vehicle_type"], 3_200) * row["max_trips_per_day"]
        for _, row in vehicles_df.iterrows()
    )

    print(f"\n  Stops ordered   : {len(stops_df)}")
    print(f"  Total KG        : {total_kg:,.1f}")
    print(f"  Vehicles avail  : {len(vehicles_df)}")
    print(f"  Fleet daily cap : {fleet_daily_cap:,} kg")

    if total_kg > fleet_daily_cap:
        print(f"\n  WARNING: Demand exceeds fleet capacity. Some stops may be unassigned.")

    print(f"\n  Load by zone:")
    zone_load = stops_df.groupby("zone").agg(stops=("kg", "count"), kg=("kg", "sum"))
    for zone, zrow in zone_load.iterrows():
        print(f"    {zone:<15} {zrow['stops']:>3} stops   {zrow['kg']:>8,.1f} kg")

    print(f"\nSolving (up to {config.SOLVER_TIME_LIMIT_SECONDS} s)...")
    routes, unassigned, stops_df = solve(stops_df, vehicles_df)

    write_excel(routes, stops_df, unassigned, config.OUTPUT_FILE)

    veh_used       = len(routes)
    stops_assigned = sum(len(r["stops"]) for r in routes)
    print(f"\nDone.")
    print(f"  Output          : {config.OUTPUT_FILE}")
    print(f"  Vehicles used   : {veh_used} / {len(vehicles_df)}")
    print(f"  Stops assigned  : {stops_assigned} / {len(stops_df)}")
    if unassigned:
        print(f"  UNASSIGNED      : {len(unassigned)} stop(s) -> see 'Unassigned' sheet")

    print(f"\n  Trip breakdown:")
    for route in routes:
        trips    = split_into_trips(route)
        trip_str = "  |  ".join(
            f"Trip {t['trip_num']}: {t['trip_kg']:.0f} kg ({len(t['stops'])} stops)"
            for t in trips
        )
        print(f"    {route['vehicle_name']:<22} {route['vehicle_type']:<8}  {trip_str}")

    print()


def build_zone_summary_from_routes(routes, original_zone_counts):
    """Build zone load summary from solved routes for territory mode output."""
    zone_summary = []
    for route in routes:
        vname    = route["vehicle_name"]
        vzone    = route["vehicle_zone"]
        stops    = route["stops"]
        n_stops  = len(stops)
        total_kg = sum(s["kg"] for s in stops)
        orig     = original_zone_counts.get(vname, 0)

        in_zone    = sum(1 for s in stops if str(s.get("zone", "")) == vzone)
        cross_zone = n_stops - in_zone

        if n_stops == 0:
            status = "NO ORDERS"
        elif n_stops > config.MAX_STOPS_PER_DAY:
            status = "OVERLOADED"
        elif n_stops <= 2:
            status = "LIGHT"
        else:
            status = "OK"

        notes = []
        if cross_zone > 0:
            notes.append(f"{cross_zone} stop(s) from other zones (solver reassigned for efficiency)")

        zone_summary.append({
            "vehicle":    vname,
            "type":       route["vehicle_type"],
            "zone":       vzone,
            "orig_stops": orig,
            "stops":      n_stops,
            "kg":         round(total_kg, 1),
            "status":     status,
            "note":       "  |  ".join(notes),
        })
    return zone_summary


def main_territory():
    print("=" * 60)
    print("  DSD Route Planner - TERRITORY MODE (zone-affinity solver)")
    print("=" * 60)

    print(f"\nLoading orders and vehicles from '{config.TODAY_FILE}'...")
    stops_df    = load_orders(config.TODAY_FILE, config.CUSTOMER_MASTER)
    vehicles_df = load_vehicles(config.TODAY_FILE)

    total_kg    = stops_df["kg"].sum()
    n_veh       = len(vehicles_df)
    n_with_zone = int((~vehicles_df["zone"].str.lower()
                       .isin(["", "nan", "float", "none"])).sum())

    print(f"\n  Stops ordered   : {len(stops_df)}")
    print(f"  Total KG        : {total_kg:,.1f}")
    print(f"  Vehicles avail  : {n_veh}")
    print(f"  Vehicles w/zone : {n_with_zone}")
    print(f"  Zone penalty    : {config.ZONE_AFFINITY_PENALTY_KM} km equivalent")

    if n_with_zone == 0:
        sys.exit(
            "\n  ERROR: No vehicles have a zone assigned in today.xlsx.\n"
            "  Run suggest_zones.bat first."
        )

    # Pre-compute how many of today's stops belong to each vehicle's zone
    vehicle_zones = {}
    for _, row in vehicles_df.iterrows():
        z = str(row["zone"]).strip()
        if z.lower() not in ("", "nan", "float", "none"):
            vehicle_zones[row["vehicle_name"]] = z

    original_zone_counts = {}
    for _, row in stops_df.iterrows():
        sz = str(row.get("zone", "")).strip()
        for vname, vzone in vehicle_zones.items():
            if vzone == sz:
                original_zone_counts[vname] = original_zone_counts.get(vname, 0) + 1
                break

    print(f"\n  Load by zone:")
    zone_load = stops_df.groupby("zone").agg(stops=("kg", "count"), kg=("kg", "sum"))
    for zone, zrow in zone_load.iterrows():
        print(f"    {zone:<22} {zrow['stops']:>3} stops   {zrow['kg']:>8,.1f} kg")

    print(f"\nSolving with zone affinity (up to {config.SOLVER_TIME_LIMIT_SECONDS} s)...")
    routes, unassigned, stops_df = solve(stops_df, vehicles_df, zone_affinity=True)

    zone_summary = build_zone_summary_from_routes(routes, original_zone_counts)

    write_excel(routes, stops_df, unassigned, config.OUTPUT_FILE, zone_summary)

    stops_assigned = sum(len(r["stops"]) for r in routes)
    overloaded = [z for z in zone_summary if z["status"] == "OVERLOADED"]
    light      = [z for z in zone_summary if z["status"] == "LIGHT"]
    no_orders  = [z for z in zone_summary if z["status"] == "NO ORDERS"]

    print(f"\nDone.")
    print(f"  Output          : {config.OUTPUT_FILE}")
    print(f"  Vehicles routed : {len(routes)} / {n_veh}")
    print(f"  Stops assigned  : {stops_assigned} / {len(stops_df)}")
    if unassigned:
        print(f"  UNASSIGNED      : {len(unassigned)} stop(s) -> see 'Unassigned' sheet")

    print(f"\n  Final stop distribution:")
    for zrow in sorted(zone_summary, key=lambda z: -z["stops"]):
        if zrow["stops"] == 0: continue
        orig  = zrow.get("orig_stops", zrow["stops"])
        delta = zrow["stops"] - orig
        arrow = f"  ({'+' if delta >= 0 else ''}{delta} solver reassigned)" if delta != 0 else ""
        print(f"    {zrow['vehicle']:<22} {zrow['zone']:<22} "
              f"{zrow['stops']:>3} stops  {zrow['kg']:>9,.1f} kg{arrow}")

    print(f"\n  Zone status summary:")
    print(f"    OK             : {sum(1 for z in zone_summary if z['status']=='OK')}")
    if overloaded:
        print(f"    OVERLOADED     : {len(overloaded)} ({', '.join(z['zone'] for z in overloaded)})")
    if light:
        print(f"    LIGHT (<3 stop): {len(light)} ({', '.join(z['zone'] for z in light)})")
    if no_orders:
        print(f"    NO ORDERS      : {len(no_orders)} vehicles idle today")
    print()
    if no_orders:
        print(f"    NO ORDERS      : {len(no_orders)} vehicles idle today")
    print()


def main():
    mode = sys.argv[1].lower() if len(sys.argv) > 1 else "optimise"
    if mode == "territory":
        main_territory()
    else:
        _main_optimise()


if __name__ == "__main__":
    main()
